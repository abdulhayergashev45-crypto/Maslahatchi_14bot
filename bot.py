"""
Maktab Maslahatchisi Bot - v6
Ikki rejimli: Maslahatchi (Admin) + O'quvchi

MASLAHATCHI:
- O'quvchilar (qo'shish, Excel import)
- To'garaklar (yaratish)
- Yutuq qo'shish
- Portfel (AI)
- MOCK test yaratish
- Hisobot (AI tahlil)

O'QUVCHI:
- Ro'yxatdan o'tish (ism + sinf → profil)
- To'garaklar (ko'rish, a'zo bo'lish)
- Mening yutuqlarim
- MOCK test yechish
- Adminga murojaat
"""

import asyncio
import logging
import os
import sqlite3
import io
from datetime import datetime

from google import genai
from google.genai import types
import openpyxl
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ContextTypes, filters, ConversationHandler
)
from telegram.constants import ParseMode

# ─── SOZLAMALAR ────────────────────────────────────────────────────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "YOUR_GEMINI_KEY")
ADMIN_IDS = list(map(int, os.getenv("ADMIN_IDS", "123456789").split(",")))

gemini_client = genai.Client(api_key=GEMINI_API_KEY)

# ─── STATES ────────────────────────────────────────────────────────────────────
(
    # Umumiy
    MAIN_MENU,
    # O'quvchi ro'yxatdan o'tish
    REG_NAME, REG_CLASS,
    # Admin — o'quvchi qo'shish
    S_NAME, S_CLASS, S_DATA,
    # Qidirish
    S_SEARCH,
    # Portfel
    PORTFOLIO,
    # To'garak
    C_NAME, C_DIR, C_RESP,
    # Yutuq
    A_STUDENT, A_TITLE, A_TYPE, A_DATE,
    # AI chat
    AI_CHAT,
    # To'garakka a'zo
    M_SEARCH, M_CLUB,
    # DTM
    DTM_FAN1, DTM_FAN2, DTM_BALL,
    # MOCK yaratish
    MOCK_NEW_TITLE, MOCK_NEW_CLASS, MOCK_NEW_FAN,
    MOCK_SAVOL, MOCK_JAVOB_TEXT,
    # MOCK yechish
    MOCK_YECHISH,
    # Murojaat
    MUROJAAT_TEXT,
) = range(28)

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO,
    handlers=[logging.FileHandler("bot.log", encoding="utf-8"), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# ─── DATABASE ──────────────────────────────────────────────────────────────────
DB_PATH = "maktab.db"

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            class_name TEXT,
            telegram_id INTEGER UNIQUE,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS student_media (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER, media_type TEXT,
            content TEXT, caption TEXT,
            added_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS clubs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, direction TEXT,
            responsible TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS club_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            club_id INTEGER, student_id INTEGER,
            joined_at TEXT DEFAULT (datetime('now')),
            UNIQUE(club_id, student_id)
        );
        CREATE TABLE IF NOT EXISTS achievements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER, title TEXT,
            achievement_type TEXT, date TEXT,
            added_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS mock_tests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT, class_name TEXT, fan TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            is_active INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS mock_questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            test_id INTEGER, question_text TEXT,
            q_type TEXT DEFAULT 'abcd',
            correct_answer TEXT,
            option_a TEXT, option_b TEXT,
            option_c TEXT, option_d TEXT,
            order_num INTEGER
        );
        CREATE TABLE IF NOT EXISTS mock_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            test_id INTEGER, student_id INTEGER,
            student_name TEXT, score INTEGER,
            total INTEGER, percentage REAL,
            answers TEXT,
            taken_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS murojaatlar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER, student_name TEXT,
            message TEXT, status TEXT DEFAULT 'yangi',
            created_at TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit(); conn.close()
    logger.info("✅ Database tayyor")

def db(): return sqlite3.connect(DB_PATH)

# ─── GEMINI AI ─────────────────────────────────────────────────────────────────
def gemini(prompt: str, system: str = "") -> str:
    try:
        config = types.GenerateContentConfig(
            system_instruction=system if system else None
        )
        response = gemini_client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
            config=config
        )
        return response.text
    except Exception as e:
        logger.error(f"Gemini: {e}")
        return f"❌ AI xatosi: {e}"

def ai_portfolio(name, cls, data):
    items = "\n".join([
        f"[{d['added_at'][:10]}] {d['media_type'].upper()}: {d['content']}"
        + (f" | {d['caption']}" if d['caption'] else "")
        for d in data
    ])
    return gemini(
        f"O'quvchi: {name} | Sinf: {cls or '—'}\nMa'lumotlar:\n{items}\n\n"
        f"Sana: {datetime.now().strftime('%Y-%m-%d')}\n\n"
        f"Quyidagi tuzilmada O'ZBEK TILIDA professional ijtimoiy portfel yozing:\n"
        f"📋 IJTIMOIY PORTFEL: {name}\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 UMUMIY MA'LUMOT\n🏆 YUTUQLAR\n🎭 TO'GARAKLAR\n"
        f"🎯 KELAJAK REJALARI\n💡 SHAXSIY SIFATLAR\n📊 MASLAHATCHI XULOSASI\n━━━━━━━━━━━━━━━━━━━━━━━━",
        "Siz tajribali maktab maslahatchisisiz."
    )

def ai_report(rtype, data):
    return gemini(
        f"{rtype} hisoboti:\n{data}\n\nProfessional hisobot, statistika va tavsiyalar bering. O'zbek tilida.",
        "Siz maktab maslahatchisisiz."
    )

def ai_guide(question, menu=""):
    return gemini(question,
        f"Siz maktab maslahatchisi botining yordamchisisiz. O'zbek tilida qisqa javob bering. Menyu: {menu}")

# ─── KLAVIATURALAR ─────────────────────────────────────────────────────────────
def admin_kb():
    return ReplyKeyboardMarkup([
        [KeyboardButton("👨‍🎓 O'quvchilar"), KeyboardButton("🏆 Yutuqlar")],
        [KeyboardButton("🎭 To'garaklar"), KeyboardButton("🎯 Kasb yo'naltirish")],
        [KeyboardButton("📋 Portfel"), KeyboardButton("📝 MOCK test")],
        [KeyboardButton("📊 Hisobot"), KeyboardButton("📨 Murojaatlar")],
        [KeyboardButton("📊 DTM Ballarni yangilash")],
    ], resize_keyboard=True, is_persistent=True)

def student_kb():
    return ReplyKeyboardMarkup([
        [KeyboardButton("🎭 To'garaklar"), KeyboardButton("🏆 Mening yutuqlarim")],
        [KeyboardButton("📝 MOCK test yechish"), KeyboardButton("📨 Adminga murojaat")],
        [KeyboardButton("🎯 OTM topish (DTM ball)"), KeyboardButton("👤 Mening profilim")],
    ], resize_keyboard=True, is_persistent=True)

MENU_ITEMS_ADMIN = {
    "👨‍🎓 O'quvchilar", "🏆 Yutuqlar", "🎭 To'garaklar",
    "🎯 Kasb yo'naltirish", "📋 Portfel", "📝 MOCK test",
    "📊 Hisobot", "📨 Murojaatlar", "📊 DTM Ballarni yangilash"
}
MENU_ITEMS_STUDENT = {
    "🎭 To'garaklar", "🏆 Mening yutuqlarim",
    "📝 MOCK test yechish", "📨 Adminga murojaat",
    "🎯 OTM topish (DTM ball)", "👤 Mening profilim"
}
MENU_ITEMS = MENU_ITEMS_ADMIN | MENU_ITEMS_STUDENT

DIR_KB = InlineKeyboardMarkup([
    [InlineKeyboardButton("🎨 Madaniyat", callback_data="cd_madaniyat"),
     InlineKeyboardButton("💻 Texnologiya", callback_data="cd_texnologiya")],
    [InlineKeyboardButton("⚽ Sport", callback_data="cd_sport"),
     InlineKeyboardButton("🎭 San'at", callback_data="cd_sanat")],
    [InlineKeyboardButton("🌿 Ekologiya", callback_data="cd_ekologiya"),
     InlineKeyboardButton("📌 Boshqa", callback_data="cd_boshqa")],
])
DIR_ICONS = {"madaniyat":"🎨","texnologiya":"💻","sport":"⚽","sanat":"🎭","ekologiya":"🌿","boshqa":"📌"}
ACH_ICONS = {"olimpiada":"🧮","sport":"⚽","sanat":"🎨","texnologiya":"💻","boshqa":"📌"}

# ─── HELPER FUNKSIYALAR ────────────────────────────────────────────────────────
def is_admin(uid): return uid in ADMIN_IDS

def get_student_by_tg(telegram_id):
    conn = db()
    row = conn.execute(
        "SELECT id, full_name, class_name FROM students WHERE telegram_id=?",
        (telegram_id,)
    ).fetchone()
    conn.close()
    return row  # (id, name, class) yoki None

async def chk_admin(update):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Faqat maslahatchi uchun.")
        return False
    return True

# ─── START ─────────────────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    context.user_data.clear()

    if is_admin(uid):
        await update.message.reply_text(
            f"🏫 *Maktab Maslahatchisi Bot* v6\n\n"
            f"Assalomu alaykum, *{update.effective_user.first_name}*! 👋\n\n"
            f"*Maslahatchi paneli* — barcha funksiyalar mavjud.",
            parse_mode=ParseMode.MARKDOWN, reply_markup=admin_kb()
        )
        return MAIN_MENU

    # O'quvchi — avval ro'yxatdan o'tganmi?
    student = get_student_by_tg(uid)
    if student:
        context.user_data["student_id"] = student[0]
        context.user_data["student_name"] = student[1]
        context.user_data["student_class"] = student[2]
        await update.message.reply_text(
            f"👋 Qaytib keldingiz, *{student[1]}*!\n"
            f"📚 Sinf: {student[2] or '—'}\n\n"
            f"Quyidagi menyudan foydalaning:",
            parse_mode=ParseMode.MARKDOWN, reply_markup=student_kb()
        )
        return MAIN_MENU

    # Yangi o'quvchi — ro'yxatdan o'tish
    await update.message.reply_text(
        "🏫 *Maktab Maslahatchisi Botiga xush kelibsiz!*\n\n"
        "Ro'yxatdan o'tish uchun *to'liq ismingizni* kiriting:\n"
        "_(Masalan: Karimov Jasur Aliyevich)_",
        parse_mode=ParseMode.MARKDOWN
    )
    return REG_NAME

# ─── O'QUVCHI RO'YXATDAN O'TISH ───────────────────────────────────────────────
async def reg_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if len(name) < 3:
        await update.message.reply_text("❗ To'liq ismingizni kiriting:")
        return REG_NAME
    context.user_data["reg_name"] = name
    await update.message.reply_text(
        f"✅ Ism: *{name}*\n\n"
        f"Sinfingizni kiriting:\n_(Masalan: 9-A, 10-B, 11-V)_",
        parse_mode=ParseMode.MARKDOWN
    )
    return REG_CLASS

async def reg_class(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cls = update.message.text.strip()
    name = context.user_data["reg_name"]
    uid = update.effective_user.id

    conn = db(); c = conn.cursor()
    # Bazada shu ism bor-yo'qligini tekshirish
    existing = c.execute(
        "SELECT id FROM students WHERE full_name=?", (name,)
    ).fetchone()

    if existing:
        # Mavjud o'quvchiga telegram_id ni bog'lash
        c.execute("UPDATE students SET telegram_id=?, class_name=? WHERE id=?",
                  (uid, cls, existing[0]))
        sid = existing[0]
    else:
        # Yangi o'quvchi qo'shish
        c.execute("INSERT INTO students (full_name, class_name, telegram_id) VALUES(?,?,?)",
                  (name, cls, uid))
        sid = c.lastrowid
    conn.commit(); conn.close()

    context.user_data["student_id"] = sid
    context.user_data["student_name"] = name
    context.user_data["student_class"] = cls

    await update.message.reply_text(
        f"✅ *Ro'yxatdan o'tdingiz!*\n\n"
        f"👤 Ism: *{name}*\n"
        f"📚 Sinf: *{cls}*\n\n"
        f"Endi botdan foydalanishingiz mumkin! 🎉",
        parse_mode=ParseMode.MARKDOWN, reply_markup=student_kb()
    )
    return MAIN_MENU

# ═══════════════════════════════════════════════════════════════
# O'QUVCHI MENYULARI
# ═══════════════════════════════════════════════════════════════


async def student_dtm_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """O'quvchi uchun DTM ball orqali OTM topish"""
    uid = update.effective_user.id
    student = get_student_by_tg(uid)

    # Oxirgi MOCK test natijasini tekshirish
    last_score_text = ""
    if student:
        conn = db()
        last_result = conn.execute(
            "SELECT score, total, percentage, taken_at FROM mock_results "
            "WHERE student_name=? ORDER BY taken_at DESC LIMIT 1",
            (student[1],)).fetchone()
        conn.close()
        if last_result:
            score, total, pct, taken = last_result
            last_score_text = (
                f"\n📊 *Oxirgi MOCK natijangiz:*\n"
                f"✅ {score}/{total} | {pct:.0f}% | 📅 {taken[:10]}\n"
                f"_(Bu ballni DTM ga moslashtiring: {score*1.4:.0f} — taxminiy)_\n"
            )

    from dtm_base import FANLAR
    kb_rows = []
    row = []
    for fan in FANLAR:
        row.append(InlineKeyboardButton(fan, callback_data=f"dtm_f1_{fan}"))
        if len(row) == 2: kb_rows.append(row); row = []
    if row: kb_rows.append(row)

    await update.message.reply_text(
        f"🎯 *DTM BALL BO'YICHA OTM TOPISH*{last_score_text}\n\n"
        f"1-fanni tanlang:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(kb_rows))
    return DTM_FAN1

async def student_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    student = get_student_by_tg(uid)
    if not student:
        await update.message.reply_text("❗ Avval /start bosing.")
        return MAIN_MENU
    sid, name, cls = student
    conn = db()
    mc = conn.execute("SELECT COUNT(*) FROM student_media WHERE student_id=?", (sid,)).fetchone()[0]
    ac = conn.execute("SELECT COUNT(*) FROM achievements WHERE student_id=?", (sid,)).fetchone()[0]
    clubs = [r[0] for r in conn.execute(
        "SELECT cl.name FROM clubs cl JOIN club_members cm ON cl.id=cm.club_id WHERE cm.student_id=?",
        (sid,)).fetchall()]
    conn.close()
    clubs_text = ', '.join(clubs) if clubs else '—'
    text = (
        f"👤 *{name}*\n"
        f"📚 Sinf: {cls or '—'}\n"
        f"📁 Ma'lumotlar: {mc} ta\n"
        f"🏆 Yutuqlar: {ac} ta\n"
        f"🎭 To'garaklar: {clubs_text}"
    )
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

async def student_my_achievements(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    student = get_student_by_tg(uid)
    if not student:
        await update.message.reply_text("❗ Avval /start bosing.")
        return MAIN_MENU
    sid, name, cls = student
    conn = db()
    achs = conn.execute(
        "SELECT title, achievement_type, date FROM achievements WHERE student_id=? ORDER BY date DESC",
        (sid,)).fetchall()
    media = conn.execute(
        "SELECT content, added_at FROM student_media WHERE student_id=? AND media_type='text' ORDER BY added_at DESC LIMIT 10",
        (sid,)).fetchall()
    conn.close()

    if not achs and not media:
        await update.message.reply_text(
            f"📭 *{name}*, hali yutuqlaringiz kiritilmagan.\n\n"
            f"Maslahatchi sizning yutuqlaringizni kiritgandan so'ng bu yerda ko'rinadi.",
            parse_mode=ParseMode.MARKDOWN
        )
        return MAIN_MENU

    text = f"🏆 *{name} — YUTUQLARIM*\n━━━━━━━━━━━━━━━\n\n"
    if achs:
        text += "🥇 *Rasmiy yutuqlar:*\n"
        for title, atype, date in achs:
            icon = ACH_ICONS.get(atype, "🏅")
            text += f"{icon} {title}\n   📅 {date or '—'}\n\n"
    if media:
        text += "📝 *Qo'shimcha ma'lumotlar:*\n"
        for content, added_at in media[:5]:
            text += f"• {content[:80]}\n"

    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

async def student_clubs_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """O'quvchi uchun to'garaklar menyusi"""
    uid = update.effective_user.id
    student = get_student_by_tg(uid)
    sid = student[0] if student else None

    conn = db()
    all_clubs = conn.execute(
        "SELECT cl.id, cl.name, cl.direction, cl.responsible, COUNT(cm.id) FROM clubs cl "
        "LEFT JOIN club_members cm ON cl.id=cm.club_id GROUP BY cl.id ORDER BY cl.direction"
    ).fetchall()

    my_club_ids = set()
    if sid:
        my_clubs = conn.execute(
            "SELECT club_id FROM club_members WHERE student_id=?", (sid,)
        ).fetchall()
        my_club_ids = {r[0] for r in my_clubs}
    conn.close()

    if not all_clubs:
        await update.message.reply_text(
            "📭 Hali to'garaklar yaratilmagan.\nMaslahatchi to'garaklarni kiritgandan so'ng ko'rinadi."
        )
        return MAIN_MENU

    text = "🎭 *TO'GARAKLAR*\n\n"
    kb_rows = []
    for cid, cname, direction, resp, cnt in all_clubs:
        icon = DIR_ICONS.get(direction, "📌")
        joined = "✅" if cid in my_club_ids else ""
        text += f"{icon} *{cname}* {joined}\n   👤 {resp or '—'} | 👥 {cnt} a'zo\n\n"
        if cid not in my_club_ids:
            kb_rows.append([InlineKeyboardButton(
                f"➕ {cname} ga a'zo bo'lish", callback_data=f"sjoin_{cid}"
            )])
        else:
            kb_rows.append([InlineKeyboardButton(
                f"✅ {cname} (a'zosiz)", callback_data=f"sleave_{cid}"
            )])

    if is_admin(uid):
        pass  # Admin uchun alohida menyu bor
    await update.message.reply_text(
        text, parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(kb_rows) if kb_rows else None
    )
    return MAIN_MENU

async def student_join_club(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    cid = int(q.data.split("_")[-1])
    uid = q.from_user.id
    student = get_student_by_tg(uid)
    if not student:
        await q.message.reply_text("❗ Avval /start bosing.")
        return MAIN_MENU
    sid, name, cls = student
    conn = db(); c = conn.cursor()
    club = conn.execute("SELECT name FROM clubs WHERE id=?", (cid,)).fetchone()
    try:
        c.execute("INSERT INTO club_members(club_id,student_id) VALUES(?,?)", (cid, sid))
        c.execute("INSERT INTO student_media(student_id,media_type,content,added_at) VALUES(?,?,?,?)",
                  (sid, "text", f"TO'GARAK: {club[0]} ga a'zo bo'ldi", datetime.now().isoformat()))
        conn.commit()
        await q.message.reply_text(
            f"✅ *{name}*, siz *{club[0]}* to'garaklariga a'zo bo'ldingiz!",
            parse_mode=ParseMode.MARKDOWN
        )
    except:
        await q.message.reply_text(f"⚠️ Siz allaqachon bu to'garakda a'zosiz.")
    conn.close()
    return MAIN_MENU

async def student_leave_club(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    cid = int(q.data.split("_")[-1])
    uid = q.from_user.id
    student = get_student_by_tg(uid)
    if not student:
        return MAIN_MENU
    sid = student[0]
    conn = db(); c = conn.cursor()
    club = conn.execute("SELECT name FROM clubs WHERE id=?", (cid,)).fetchone()
    c.execute("DELETE FROM club_members WHERE club_id=? AND student_id=?", (cid, sid))
    conn.commit(); conn.close()
    await q.message.reply_text(f"✅ *{club[0]}* to'garakdan chiqdingiz.", parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

# ─── MUROJAAT ──────────────────────────────────────────────────────────────────
async def murojaat_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if is_admin(uid):
        return await murojaatlar_admin(update, context)
    student = get_student_by_tg(uid)
    if not student:
        await update.message.reply_text("❗ Avval /start bosing.")
        return MAIN_MENU
    await update.message.reply_text(
        "📨 *Adminga murojaat*\n\n"
        "Xabaringizni yozing — maslahatchi ko'radi va javob beradi:\n\n"
        "_(Savolingiz, muammoingiz, taklifingiz)_",
        parse_mode=ParseMode.MARKDOWN
    )
    return MUROJAAT_TEXT

async def murojaat_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS:
        return await main_router(update, context)
    uid = update.effective_user.id
    student = get_student_by_tg(uid)
    if not student:
        return MAIN_MENU
    sid, name, cls = student
    text = update.message.text.strip()
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO murojaatlar(student_id,student_name,message) VALUES(?,?,?)",
              (sid, name, text))
    conn.commit(); conn.close()

    # Adminga xabar yuborish
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                admin_id,
                f"📨 *Yangi murojaat!*\n\n"
                f"👤 {name} | 📚 {cls or '—'}\n"
                f"📝 {text}\n\n"
                f"_Javob berish uchun o'quvchiga to'g'ridan to'g'ri yozing_",
                parse_mode=ParseMode.MARKDOWN
            )
        except Exception as e:
            logger.warning(f"Admin ga xabar yuborishda xato: {e}")

    await update.message.reply_text(
        "✅ *Murojaat yuborildi!*\n\nMaslahatchi tez orada ko'radi.",
        parse_mode=ParseMode.MARKDOWN, reply_markup=student_kb()
    )
    return MAIN_MENU

async def murojaatlar_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = db()
    rows = conn.execute(
        "SELECT id, student_name, message, status, created_at FROM murojaatlar ORDER BY created_at DESC LIMIT 20"
    ).fetchall()
    conn.close()
    if not rows:
        await update.message.reply_text("📭 Hali murojaatlar yo'q.")
        return MAIN_MENU
    text = "📨 *MUROJAATLAR:*\n\n"
    for mid, sname, msg, status, created in rows:
        icon = "🆕" if status == "yangi" else "✅"
        text += f"{icon} *{sname}*\n   {msg[:100]}\n   📅 {created[:16]}\n\n"
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

# ═══════════════════════════════════════════════════════════════
# ADMIN MENYULARI
# ═══════════════════════════════════════════════════════════════

# ─── O'QUVCHILAR (ADMIN) ───────────────────────────────────────
async def admin_students_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Yangi o'quvchi qo'shish", callback_data="add_student")],
        [InlineKeyboardButton("📥 Excel yuklash (massa)", callback_data="excel_help")],
        [InlineKeyboardButton("🔍 O'quvchi qidirish", callback_data="srch_student")],
        [InlineKeyboardButton("📋 Barcha o'quvchilar", callback_data="list_students")],
    ])
    await update.message.reply_text(
        "👨‍🎓 *O'QUVCHILAR BOSHQARUVI*",
        parse_mode=ParseMode.MARKDOWN, reply_markup=kb
    )
    return MAIN_MENU

async def cb_add_student(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await update.callback_query.message.reply_text(
        "➕ O'quvchining *to'liq ismini* kiriting:",
        parse_mode=ParseMode.MARKDOWN)
    return S_NAME

async def s_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    context.user_data["sname"] = update.message.text.strip()
    await update.message.reply_text(
        f"✅ Ism: *{context.user_data['sname']}*\n\nSinfini kiriting _(9-A, 10-B)_:",
        parse_mode=ParseMode.MARKDOWN)
    return S_CLASS

async def s_class(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    name = context.user_data["sname"]; cls = update.message.text.strip()
    conn = db(); c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO students (full_name,class_name) VALUES(?,?)", (name, cls))
    sid = c.lastrowid; conn.commit(); conn.close()
    await update.message.reply_text(
        f"✅ *{name}* ({cls}) qo'shildi! 🆔 ID: `{sid}`\n\n"
        f"Ma'lumot qo'shish: `/add_media {sid}`",
        parse_mode=ParseMode.MARKDOWN, reply_markup=admin_kb())
    return MAIN_MENU

async def cb_srch_student(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await update.callback_query.message.reply_text("🔍 Ism yoki sinf kiriting:")
    return S_SEARCH

async def s_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    search = update.message.text.strip()
    conn = db()
    rows = conn.execute(
        "SELECT id,full_name,class_name FROM students WHERE full_name LIKE ? OR class_name LIKE ? LIMIT 10",
        (f"%{search}%", f"%{search}%")).fetchall()
    conn.close()
    if not rows:
        await update.message.reply_text("❌ Topilmadi.", reply_markup=admin_kb())
        return MAIN_MENU
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(f"👤 {n} ({cls or '—'})", callback_data=f"sp_{sid}")]
                                for sid, n, cls in rows])
    await update.message.reply_text(f"🔍 *{len(rows)} natija:*",
                                     parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def cb_list_students(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    conn = db()
    rows = conn.execute(
        "SELECT id,full_name,class_name,telegram_id FROM students ORDER BY class_name,full_name LIMIT 50"
    ).fetchall()
    conn.close()
    if not rows:
        await update.callback_query.message.reply_text("📭 O'quvchilar yo'q.")
        return MAIN_MENU
    text = "📋 *BARCHA O'QUVCHILAR:*\n\n"
    for sid, n, cls, tgid in rows:
        online = "🟢" if tgid else "⚪"
        text += f"{online} `{sid}` | {n} | {cls or '—'}\n"
    text += "\n🟢 = Bot orqali ro'yxatdan o'tgan"
    await update.callback_query.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

async def cb_student_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    sid = int(q.data.split("_")[-1])
    conn = db()
    s = conn.execute("SELECT full_name,class_name,telegram_id FROM students WHERE id=?", (sid,)).fetchone()
    if not s: conn.close(); return MAIN_MENU
    name, cls, tgid = s
    mc = conn.execute("SELECT COUNT(*) FROM student_media WHERE student_id=?", (sid,)).fetchone()[0]
    ac = conn.execute("SELECT COUNT(*) FROM achievements WHERE student_id=?", (sid,)).fetchone()[0]
    clubs = [r[0] for r in conn.execute(
        "SELECT cl.name FROM clubs cl JOIN club_members cm ON cl.id=cm.club_id WHERE cm.student_id=?",
        (sid,)).fetchall()]
    conn.close()
    clubs_text = ', '.join(clubs) if clubs else '—'
    online = '🟢 Botda royxatdan otgan' if tgid else '⚪ Royxatdan otmagan'
    text = (f"👤 *{name}* | 📚 {cls or chr(8212)}\n"
            f"{online}\n"
            f"📁 {mc} ta | 🏆 {ac} ta\n"
            f"🎭 {clubs_text}")
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Portfel", callback_data=f"pf_{sid}"),
         InlineKeyboardButton("📁 Ma\'lumot", callback_data=f"am_{sid}")],
        [InlineKeyboardButton("🏆 Yutuq qo\'sh", callback_data=f"aa_{sid}")],
    ])
    await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

# ─── MEDIA ─────────────────────────────────────────────────────
async def add_media_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await chk_admin(update): return
    if not context.args:
        await update.message.reply_text("❗ `/add_media <ID>`", parse_mode=ParseMode.MARKDOWN)
        return
    try: sid = int(context.args[0])
    except: await update.message.reply_text("❗ ID raqam bo'lishi kerak."); return
    conn = db()
    s = conn.execute("SELECT full_name FROM students WHERE id=?", (sid,)).fetchone()
    conn.close()
    if not s: await update.message.reply_text(f"❌ ID={sid} topilmadi."); return
    context.user_data["tsid"] = sid
    context.user_data["tsname"] = s[0]
    await update.message.reply_text(
        f"📁 *{s[0]}* uchun yuboring:\n📝 Matn | 🖼 Rasm | 🎥 Video\n\n✅ Tugash: /done",
        parse_mode=ParseMode.MARKDOWN)
    return S_DATA

async def receive_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    sid = context.user_data.get("tsid")
    if not sid: return MAIN_MENU
    conn = db(); c = conn.cursor(); now = datetime.now().isoformat()
    m = update.message
    if m.text and not m.text.startswith("/"):
        c.execute("INSERT INTO student_media(student_id,media_type,content,added_at)VALUES(?,?,?,?)",
                  (sid,"text",m.text,now))
        await m.reply_text("✅ Matn saqlandi!")
    elif m.photo:
        c.execute("INSERT INTO student_media(student_id,media_type,content,caption,added_at)VALUES(?,?,?,?,?)",
                  (sid,"photo",m.photo[-1].file_id,m.caption or "",now))
        await m.reply_text("✅ Rasm saqlandi!")
    elif m.video:
        c.execute("INSERT INTO student_media(student_id,media_type,content,caption,added_at)VALUES(?,?,?,?,?)",
                  (sid,"video",m.video.file_id,m.caption or "",now))
        await m.reply_text("✅ Video saqlandi!")
    elif m.document and not (m.document.file_name or "").endswith(".xlsx"):
        c.execute("INSERT INTO student_media(student_id,media_type,content,caption,added_at)VALUES(?,?,?,?,?)",
                  (sid,"document",m.document.file_id,m.caption or "",now))
        await m.reply_text("✅ Hujjat saqlandi!")
    conn.commit(); conn.close()
    return S_DATA

async def done_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = context.user_data.pop("tsname", "O'quvchi")
    context.user_data.pop("tsid", None)
    context.user_data.pop("ai_mode", None)
    kb = admin_kb() if is_admin(update.effective_user.id) else student_kb()
    await update.message.reply_text(
        f"✅ *{name}* uchun saqlash yakunlandi!",
        parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

# ─── YUTUQLAR (ADMIN) ──────────────────────────────────────────
async def admin_achievements_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Yutuq qo'shish", callback_data="add_ach")],
        [InlineKeyboardButton("📊 Barcha yutuqlar", callback_data="list_ach")],
    ])
    await update.message.reply_text("🏆 *YUTUQ VA OLIMPIADALAR*",
                                     parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def cb_add_ach(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await update.callback_query.message.reply_text("🏆 O'quvchi ismini kiriting:")
    return A_STUDENT

async def a_student(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    conn = db()
    rows = conn.execute(
        "SELECT id,full_name,class_name FROM students WHERE full_name LIKE ? LIMIT 5",
        (f"%{update.message.text.strip()}%",)).fetchall()
    conn.close()
    if not rows:
        await update.message.reply_text("❌ Topilmadi:"); return A_STUDENT
    if len(rows) == 1:
        context.user_data.update({"asid": rows[0][0], "asname": rows[0][1]})
        await update.message.reply_text(
            f"✅ *{rows[0][1]}*\n\nYutuq nomini kiriting:", parse_mode=ParseMode.MARKDOWN)
        return A_TITLE
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(f"{n} ({cls or '—'})", callback_data=f"asel_{sid}")]
                                for sid, n, cls in rows])
    await update.message.reply_text("Qaysi o'quvchi?", reply_markup=kb)
    return A_STUDENT

async def cb_asel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    sid = int(q.data.split("_")[-1])
    row = db().execute("SELECT full_name FROM students WHERE id=?", (sid,)).fetchone()
    context.user_data.update({"asid": sid, "asname": row[0]})
    await q.message.reply_text(f"✅ *{row[0]}*\n\nYutuq nomini kiriting:", parse_mode=ParseMode.MARKDOWN)
    return A_TITLE

async def a_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    context.user_data["atitle"] = update.message.text.strip()
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🧮 Olimpiada", callback_data="at_olimpiada"),
         InlineKeyboardButton("⚽ Sport", callback_data="at_sport")],
        [InlineKeyboardButton("🎨 San'at", callback_data="at_sanat"),
         InlineKeyboardButton("💻 Texnologiya", callback_data="at_texnologiya")],
        [InlineKeyboardButton("📌 Boshqa", callback_data="at_boshqa")],
    ])
    await update.message.reply_text("Tur:", reply_markup=kb)
    return A_TYPE

async def cb_atype(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data["atype"] = q.data.replace("at_", "")
    await q.message.reply_text("📅 Sana _(2024-03-15)_:", parse_mode=ParseMode.MARKDOWN)
    return A_DATE

async def a_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    date = update.message.text.strip()
    sid = context.user_data["asid"]; name = context.user_data["asname"]
    title = context.user_data["atitle"]; atype = context.user_data["atype"]
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO achievements(student_id,title,achievement_type,date)VALUES(?,?,?,?)",
              (sid,title,atype,date))
    c.execute("INSERT INTO student_media(student_id,media_type,content,added_at)VALUES(?,?,?,?)",
              (sid,"text",f"YUTUQ: {title} ({atype}) — {date}",datetime.now().isoformat()))
    conn.commit(); conn.close()

    # O'quvchiga xabar yuborish
    s_tg = db().execute("SELECT telegram_id FROM students WHERE id=?", (sid,)).fetchone()
    if s_tg and s_tg[0]:
        try:
            await context.bot.send_message(
                s_tg[0],
                f"🏆 *Tabriklaymiz, {name}!*\n\n"
                f"Sizning yangi yutuqingiz kiritildi:\n"
                f"🥇 {title}\n📌 {atype} | 📅 {date}",
                parse_mode=ParseMode.MARKDOWN
            )
        except: pass

    await update.message.reply_text(
        f"✅ *{name}* yutuqi saqlandi!\n🏆 {title}",
        parse_mode=ParseMode.MARKDOWN, reply_markup=admin_kb())
    return MAIN_MENU

async def cb_list_ach(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    rows = conn.execute(
        "SELECT s.full_name,s.class_name,a.title,a.achievement_type,a.date "
        "FROM achievements a JOIN students s ON a.student_id=s.id "
        "ORDER BY a.added_at DESC LIMIT 20").fetchall()
    conn.close()
    if not rows: await q.message.reply_text("📭 Yutuqlar yo'q."); return MAIN_MENU
    text = "🏆 *SO'NGGI YUTUQLAR:*\n\n"
    for n, cls, title, atype, date in rows:
        text += f"{ACH_ICONS.get(atype,'🏅')} *{n}* ({cls or '—'})\n   {title} — {date or '—'}\n\n"
    await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

# ─── TO'GARAKLAR (ADMIN) ───────────────────────────────────────
async def admin_clubs_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ To'garak qo'shish", callback_data="add_club")],
        [InlineKeyboardButton("📋 To'garaklar ro'yxati", callback_data="list_clubs")],
        [InlineKeyboardButton("👥 A'zolar ro'yxati", callback_data="club_members_list")],
    ])
    await update.message.reply_text(
        "🎭 *TO'GARAKLAR*\n\nTez qo'shish: `/add_togarak Robototexnika`",
        parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def add_togarak_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await chk_admin(update): return
    if not context.args:
        await update.message.reply_text("❗ `/add_togarak <nom>`", parse_mode=ParseMode.MARKDOWN)
        return
    context.user_data["cname"] = " ".join(context.args)
    await update.message.reply_text(
        f"✅ To'garak: *{context.user_data['cname']}*\n\nYo'nalish:",
        parse_mode=ParseMode.MARKDOWN, reply_markup=DIR_KB)
    return C_DIR

async def cb_add_club(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await update.callback_query.message.reply_text("➕ To'garak nomini kiriting:")
    return C_NAME

async def c_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    context.user_data["cname"] = update.message.text.strip()
    await update.message.reply_text(
        f"✅ Nom: *{context.user_data['cname']}*\n\nYo'nalish:",
        parse_mode=ParseMode.MARKDOWN, reply_markup=DIR_KB)
    return C_DIR

async def cb_cdir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data["cdir"] = q.data.replace("cd_", "")
    await q.message.reply_text("👤 Mas'ul o'qituvchi ismini kiriting:")
    return C_RESP

async def c_resp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    name = context.user_data["cname"]; direction = context.user_data["cdir"]
    resp = update.message.text.strip()
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO clubs(name,direction,responsible)VALUES(?,?,?)", (name,direction,resp))
    conn.commit(); conn.close()
    await update.message.reply_text(
        f"✅ *{DIR_ICONS.get(direction,'📌')} {name}* qo'shildi!\n👤 {resp}",
        parse_mode=ParseMode.MARKDOWN, reply_markup=admin_kb())
    return MAIN_MENU

async def cb_list_clubs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    rows = conn.execute(
        "SELECT cl.name,cl.direction,cl.responsible,COUNT(cm.id) FROM clubs cl "
        "LEFT JOIN club_members cm ON cl.id=cm.club_id GROUP BY cl.id ORDER BY cl.direction").fetchall()
    conn.close()
    if not rows: await q.message.reply_text("📭 To'garaklar yo'q."); return MAIN_MENU
    text = "🎭 *TO'GARAKLAR:*\n\n"; cur = None
    for n, d, resp, cnt in rows:
        if d != cur:
            text += f"\n{DIR_ICONS.get(d,'📌')} *{(d or 'Boshqa').upper()}*\n"; cur = d
        text += f"  • {n} | 👤 {resp or '—'} | 👥 {cnt}\n"
    await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

async def cb_club_members(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    clubs = conn.execute("SELECT id,name FROM clubs").fetchall()
    conn.close()
    if not clubs: await q.message.reply_text("📭 To'garaklar yo'q."); return MAIN_MENU
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(cn, callback_data=f"cmemb_{cid}")]
                                for cid, cn in clubs])
    await q.message.reply_text("Qaysi to'garak a'zolarini ko'rasiz?", reply_markup=kb)
    return MAIN_MENU

async def cb_show_members(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    cid = int(q.data.split("_")[-1])
    conn = db()
    club = conn.execute("SELECT name FROM clubs WHERE id=?", (cid,)).fetchone()
    members = conn.execute(
        "SELECT s.full_name, s.class_name FROM students s "
        "JOIN club_members cm ON s.id=cm.student_id WHERE cm.club_id=?", (cid,)).fetchall()
    conn.close()
    if not members:
        await q.message.reply_text(f"📭 *{club[0]}* da hali a'zolar yo'q.", parse_mode=ParseMode.MARKDOWN)
        return MAIN_MENU
    text = f"👥 *{club[0]} — A'ZOLAR:*\n\n"
    for i, (name, cls) in enumerate(members, 1):
        text += f"{i}. {name} | {cls or '—'}\n"
    await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

# ─── PORTFEL ───────────────────────────────────────────────────
async def admin_portfolio_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 *PORTFEL YARATISH*\n\nO'quvchi ismi yoki ID kiriting:",
        parse_mode=ParseMode.MARKDOWN)
    return PORTFOLIO

async def portfolio_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text in MENU_ITEMS: return await handle_menu(update, context)
    conn = db()
    if text.isdigit():
        s = conn.execute("SELECT id,full_name,class_name FROM students WHERE id=?", (int(text),)).fetchone()
    else:
        s = conn.execute("SELECT id,full_name,class_name FROM students WHERE full_name LIKE ? LIMIT 1",
                         (f"%{text}%",)).fetchone()
    if not s:
        await update.message.reply_text("❌ Topilmadi. Qayta kiriting:")
        conn.close(); return PORTFOLIO
    sid, sname, cls = s
    media = conn.execute(
        "SELECT media_type,content,caption,added_at FROM student_media WHERE student_id=? ORDER BY added_at",
        (sid,)).fetchall()
    conn.close()
    if not media:
        await update.message.reply_text(
            f"⚠️ *{sname}* haqida ma'lumot yo'q.\n`/add_media {sid}`",
            parse_mode=ParseMode.MARKDOWN)
        return MAIN_MENU
    msg = await update.message.reply_text(f"⏳ *{sname}* portfeli...", parse_mode=ParseMode.MARKDOWN)
    data = [{"media_type":r[0],"content":r[1],"caption":r[2],"added_at":r[3]} for r in media]
    pt = ai_portfolio(sname, cls, data)
    await msg.delete()
    await update.message.reply_text(pt, reply_markup=admin_kb())
    return MAIN_MENU

# ─── MOCK TEST (ADMIN) ─────────────────────────────────────────
async def admin_mock_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Yangi test yaratish", callback_data="mock_create")],
        [InlineKeyboardButton("📋 Testlar ro'yxati", callback_data="mock_list")],
        [InlineKeyboardButton("📊 Natijalar", callback_data="mock_results_menu")],
    ])
    await update.message.reply_text("📝 *MOCK IMTIHON*", parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def mock_create(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data["mock"] = {"questions": []}
    await q.message.reply_text("📝 Test sarlavhasini kiriting:")
    return MOCK_NEW_TITLE

async def mock_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    context.user_data["mock"]["title"] = update.message.text.strip()
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(c, callback_data=f"mcls_{c}") for c in ["9-sinf","10-sinf","11-sinf"]],
        [InlineKeyboardButton("Barcha sinflar", callback_data="mcls_all")],
    ])
    await update.message.reply_text("Qaysi sinf?", reply_markup=kb)
    return MOCK_NEW_CLASS

async def mock_class(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data["mock"]["class"] = q.data.replace("mcls_", "")
    await q.message.reply_text("Fan nomini kiriting:")
    return MOCK_NEW_FAN

async def mock_fan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    context.user_data["mock"]["fan"] = update.message.text.strip()
    await _ask_next_q(update.message, context)
    return MOCK_SAVOL

async def _ask_next_q(message, context):
    qnum = len(context.user_data["mock"]["questions"]) + 1
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔤 A/B/C/D", callback_data="mqtype_abcd"),
         InlineKeyboardButton("✅ Ha/Yo'q", callback_data="mqtype_bool")],
        [InlineKeyboardButton("💾 Testni saqlash", callback_data="mock_save")],
    ])
    await message.reply_text(
        f"❓ *{qnum}-savol* qo'shish\n_(Hozircha {qnum-1} ta savol)_\n\nSavol turi:",
        parse_mode=ParseMode.MARKDOWN, reply_markup=kb)

async def mock_qtype(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data["mock"]["cur_qtype"] = q.data.replace("mqtype_", "")
    context.user_data["mock"]["cur_step"] = "question"
    await q.message.reply_text("✏️ Savol matnini kiriting:")
    return MOCK_JAVOB_TEXT

async def mock_qtext(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    step = context.user_data["mock"].get("cur_step")
    qtype = context.user_data["mock"]["cur_qtype"]
    if step == "question":
        context.user_data["mock"]["cur_q"] = {"text": update.message.text.strip(), "type": qtype, "options": {}}
        if qtype == "abcd":
            context.user_data["mock"]["cur_step"] = "A"
            await update.message.reply_text("A variantini kiriting:")
        else:
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Ha", callback_data="mjavc_ha"),
                InlineKeyboardButton("❌ Yo'q", callback_data="mjavc_yoq")
            ]])
            await update.message.reply_text("To'g'ri javob:", reply_markup=kb)
        return MOCK_JAVOB_TEXT
    opts = context.user_data["mock"]["cur_q"]["options"]
    next_steps = {"A": "B", "B": "C", "C": "D"}
    if step in next_steps:
        opts[step] = update.message.text.strip()
        nxt = next_steps[step]
        context.user_data["mock"]["cur_step"] = nxt
        await update.message.reply_text(f"{nxt} variantini kiriting:")
    elif step == "D":
        opts["D"] = update.message.text.strip()
        context.user_data["mock"]["cur_step"] = "correct"
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("A", callback_data="mjavc_A"),
            InlineKeyboardButton("B", callback_data="mjavc_B"),
            InlineKeyboardButton("C", callback_data="mjavc_C"),
            InlineKeyboardButton("D", callback_data="mjavc_D"),
        ]])
        await update.message.reply_text("To'g'ri javob:", reply_markup=kb)
    return MOCK_JAVOB_TEXT

async def mock_correct(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    correct = q.data.replace("mjavc_", "")
    cur_q = context.user_data["mock"]["cur_q"]
    cur_q["correct"] = correct
    context.user_data["mock"]["questions"].append(cur_q)
    await q.message.reply_text(f"✅ Savol saqlandi! Javob: *{correct}*", parse_mode=ParseMode.MARKDOWN)
    await _ask_next_q(q.message, context)
    return MOCK_SAVOL

async def mock_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    mock = context.user_data.get("mock", {})
    qs = mock.get("questions", [])
    if not qs: await q.message.reply_text("❗ Kamida 1 savol kiriting!"); return MOCK_SAVOL
    conn = db(); c = conn.cursor()
    c.execute("INSERT INTO mock_tests(title,class_name,fan)VALUES(?,?,?)",
              (mock["title"], mock["class"], mock["fan"]))
    tid = c.lastrowid
    for i, sq in enumerate(qs, 1):
        opts = sq.get("options", {})
        c.execute("INSERT INTO mock_questions(test_id,question_text,q_type,correct_answer,option_a,option_b,option_c,option_d,order_num)VALUES(?,?,?,?,?,?,?,?,?)",
                  (tid, sq["text"], sq["type"], sq["correct"],
                   opts.get("A"), opts.get("B"), opts.get("C"), opts.get("D"), i))
    conn.commit(); conn.close()
    context.user_data.pop("mock", None)
    await q.message.reply_text(
        f"✅ *Test saqlandi!*\n📝 {mock['title']}\n"
        f"📚 {mock['fan']} | 🎓 {mock['class']}\n"
        f"❓ {len(qs)} ta savol\n🆔 ID: `{tid}`\n\n"
        f"O'quvchilar uchun: `/mock_yech {tid}`",
        parse_mode=ParseMode.MARKDOWN, reply_markup=admin_kb())
    return MAIN_MENU

async def mock_list_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    tests = conn.execute(
        "SELECT id,title,class_name,fan,created_at FROM mock_tests ORDER BY created_at DESC LIMIT 10"
    ).fetchall()
    conn.close()
    if not tests: await q.message.reply_text("📭 Testlar yo'q."); return MAIN_MENU
    text = "📋 *MOCK TESTLAR:*\n\n"
    for tid, title, cls, fan, created in tests:
        text += f"🆔 `{tid}` | *{title}*\n   📚 {fan} | {cls} | {created[:10]}\n\n"
    text += "Yechish uchun: `/mock_yech <ID>`"
    await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

async def mock_results_menu_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    tests = conn.execute("SELECT id,title FROM mock_tests ORDER BY created_at DESC LIMIT 10").fetchall()
    conn.close()
    if not tests: await q.message.reply_text("📭 Testlar yo'q."); return MAIN_MENU
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(f"📊 {t[1][:30]}", callback_data=f"mres_{t[0]}")]
                                for t in tests])
    await q.message.reply_text("Qaysi test natijasi?", reply_markup=kb)
    return MAIN_MENU

async def mock_result_show(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    tid = int(q.data.replace("mres_", ""))
    conn = db()
    test = conn.execute("SELECT title,class_name,fan FROM mock_tests WHERE id=?", (tid,)).fetchone()
    results = conn.execute(
        "SELECT student_name,score,total,percentage,taken_at FROM mock_results WHERE test_id=? ORDER BY percentage DESC",
        (tid,)).fetchall()
    conn.close()
    if not results: await q.message.reply_text("📭 Natijalar yo'q."); return MAIN_MENU
    avg = sum(r[3] for r in results) / len(results)
    text = (f"📊 *{test[0]}*\n📚 {test[2]} | 🎓 {test[1]}\n"
            f"👥 {len(results)} ta | 📈 O'rtacha: {avg:.1f}%\n━━━━━━━━━━━━━━━\n\n")
    for i, (name, score, total, pct, taken) in enumerate(results, 1):
        medal = "🥇" if i==1 else "🥈" if i==2 else "🥉" if i==3 else f"{i}."
        bar = "█"*int(pct/10) + "░"*(10-int(pct/10))
        text += f"{medal} *{name}* — {score}/{total} | {pct:.0f}%\n   |{bar}|\n\n"
    await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

# ─── MOCK YECHISH (O'QUVCHI + ADMIN) ──────────────────────────
async def mock_yech_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("❗ `/mock_yech <ID>`", parse_mode=ParseMode.MARKDOWN)
        return
    try: tid = int(context.args[0])
    except: await update.message.reply_text("❗ ID raqam bo'lishi kerak."); return
    conn = db()
    test = conn.execute("SELECT id,title,class_name,fan FROM mock_tests WHERE id=?", (tid,)).fetchone()
    if not test: await update.message.reply_text(f"❌ Test ID={tid} topilmadi."); conn.close(); return
    qs = conn.execute(
        "SELECT id,question_text,q_type,option_a,option_b,option_c,option_d FROM mock_questions WHERE test_id=? ORDER BY order_num",
        (tid,)).fetchall()
    conn.close()
    if not qs: await update.message.reply_text("❌ Bu testda savollar yo'q."); return

    # O'quvchi ma'lumoti
    uid = update.effective_user.id
    student = get_student_by_tg(uid)
    if student:
        sname = student[1]
    else:
        sname = update.effective_user.first_name or "Noma'lum"

    context.user_data["mock_yech"] = {
        "test_id": tid, "test_title": test[1],
        "questions": qs, "current": 0,
        "answers": [], "student_name": sname, "step": "quiz"
    }
    await update.message.reply_text(
        f"📝 *{test[1]}*\n📚 {test[3]} | 🎓 {test[2]}\n❓ {len(qs)} ta savol\n\n"
        f"Tayyor bo'lsangiz, boshlaylik!",
        parse_mode=ParseMode.MARKDOWN)
    await _send_mock_q(update.message, context)
    return MOCK_YECHISH

async def student_mock_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """O'quvchi uchun mock test menyusi"""
    conn = db()
    tests = conn.execute(
        "SELECT id,title,fan,class_name FROM mock_tests WHERE is_active=1 ORDER BY created_at DESC LIMIT 10"
    ).fetchall()
    conn.close()
    if not tests:
        await update.message.reply_text("📭 Hozircha testlar yo'q.\nMaslahatchi test yaratgandan so'ng ko'rinadi.")
        return MAIN_MENU
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(
        f"📝 {t[1]} ({t[2]})", callback_data=f"start_test_{t[0]}")] for t in tests])
    await update.message.reply_text(
        "📝 *MOCK TESTLAR*\n\nQaysi testni yechmoqchisiz?",
        parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def cb_start_test(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    tid = int(q.data.split("_")[-1])
    conn = db()
    test = conn.execute("SELECT id,title,class_name,fan FROM mock_tests WHERE id=?", (tid,)).fetchone()
    qs = conn.execute(
        "SELECT id,question_text,q_type,option_a,option_b,option_c,option_d FROM mock_questions WHERE test_id=? ORDER BY order_num",
        (tid,)).fetchall()
    conn.close()
    if not qs: await q.message.reply_text("❌ Savollar yo'q."); return MAIN_MENU

    uid = q.from_user.id
    student = get_student_by_tg(uid)
    sname = student[1] if student else q.from_user.first_name

    context.user_data["mock_yech"] = {
        "test_id": tid, "test_title": test[1],
        "questions": qs, "current": 0,
        "answers": [], "student_name": sname, "step": "quiz"
    }
    await _send_mock_q(q.message, context)
    return MOCK_YECHISH

async def _send_mock_q(message, context):
    my = context.user_data["mock_yech"]
    idx = my["current"]
    qs = my["questions"]
    if idx >= len(qs):
        await _finish_mock(message, context)
        return
    qid, qtext, qtype, oa, ob, oc, od = qs[idx]
    qnum = idx + 1; total = len(qs)
    if qtype == "abcd":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"A) {oa}", callback_data="mq_A"),
             InlineKeyboardButton(f"B) {ob}", callback_data="mq_B")],
            [InlineKeyboardButton(f"C) {oc}", callback_data="mq_C"),
             InlineKeyboardButton(f"D) {od}", callback_data="mq_D")],
        ])
    else:
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Ha", callback_data="mq_ha"),
            InlineKeyboardButton("❌ Yo'q", callback_data="mq_yoq"),
        ]])
    progress = "▓"*idx + "░"*(total-idx)
    await message.reply_text(
        f"❓ *{qnum}/{total}*\n{progress}\n\n{qtext}",
        parse_mode=ParseMode.MARKDOWN, reply_markup=kb)

async def mock_answer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    if not q.data.startswith("mq_"): return MOCK_YECHISH
    my = context.user_data.get("mock_yech")
    if not my: return MAIN_MENU
    my["answers"].append(q.data.replace("mq_", ""))
    my["current"] += 1
    await _send_mock_q(q.message, context)
    return MOCK_YECHISH


async def mock_to_dtm_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """MOCK natijasidan DTM OTM topishga o'tish"""
    q = update.callback_query; await q.answer()
    ball = int(q.data.replace("mock_to_dtm_", ""))
    context.user_data["mock_dtm_ball"] = ball

    from dtm_base import FANLAR
    kb_rows = []
    row = []
    for fan in FANLAR:
        row.append(InlineKeyboardButton(fan, callback_data=f"mdt_f1_{fan}"))
        if len(row) == 2: kb_rows.append(row); row = []
    if row: kb_rows.append(row)

    await q.message.reply_text(
        f"🎯 *Ball: {ball}*\n\n"
        f"Bu ballga mos OTMlarni topish uchun\n*1-fanni* tanlang:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(kb_rows))

async def mock_dtm_fan1(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    fan1 = q.data.replace("mdt_f1_", "")
    context.user_data["dtm_fan1"] = fan1
    from dtm_base import FANLAR
    kb_rows = []; row = []
    for fan in FANLAR:
        if fan == fan1: continue
        row.append(InlineKeyboardButton(fan, callback_data=f"mdt_f2_{fan}"))
        if len(row) == 2: kb_rows.append(row); row = []
    if row: kb_rows.append(row)
    await q.message.reply_text(
        f"✅ 1-fan: *{fan1}*\n\n2-fanni tanlang:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(kb_rows))

async def mock_dtm_fan2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    fan2 = q.data.replace("mdt_f2_", "")
    fan1 = context.user_data["dtm_fan1"]
    ball = context.user_data.get("mock_dtm_ball", 140)

    msg = await q.message.reply_text(
        f"🔍 *{fan1} + {fan2} | {ball} ball*\nQidirilmoqda...")

    from dtm_base import find_universities
    natija = find_universities(fan1, fan2, ball)
    grant_list = natija["grant"]
    kontrakt_list = natija["kontrakt"]

    uid = q.from_user.id
    student = get_student_by_tg(uid)
    sname = student[1] if student else "O'quvchi"

    if not grant_list and not kontrakt_list:
        await msg.edit_text(
            f"😔 *{sname}*, {ball} ball bilan {fan1}+{fan2} bo'yicha\n"
            f"hozircha mos yo'nalish topilmadi.\n\n"
            f"💡 Ballingizni oshiring yoki boshqa fan sinang.",
            parse_mode=ParseMode.MARKDOWN)
        return

    text = (f"🎯 *{sname} — OTM TAVSIYALARI*\n"
            f"📚 {fan1} + {fan2} | 🏆 {ball} ball\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\n\n")

    if grant_list:
        text += f"🎓 *GRANT OLISH MUMKIN ({len(grant_list)} ta):*\n\n"
        for i, r in enumerate(grant_list, 1):
            text += (f"{i}. 🏫 *{r['otm']}*\n"
                     f"   📚 {r['yonalish']}\n"
                     f"   🏙 {r['city']} | 🎓 {r['grant_ball']} ball kerak\n"
                     f"   ⏱ {r['muddat']} yil | 🔗 {r['site']}\n\n")

    if kontrakt_list:
        text += f"💳 *KONTRAKT ASOSIDA ({len(kontrakt_list)} ta):*\n\n"
        for i, r in enumerate(kontrakt_list, 1):
            text += (f"{i}. 🏫 *{r['otm']}*\n"
                     f"   📚 {r['yonalish']}\n"
                     f"   🏙 {r['city']} | 💳 {r['kontrakt_ball']} ball\n"
                     f"   💰 {r['narx']} mln/yil | ⏱ {r['muddat']} yil\n\n")

    text += f"━━━━━━━━━━━━━━━━━━━━━━━\n💡 _DTM ball {ball} — taxminiy hisob_"

    kb = admin_kb() if is_admin(uid) else student_kb()
    if len(text) > 4000:
        await msg.edit_text(text[:4000], parse_mode=ParseMode.MARKDOWN)
        await q.message.reply_text(text[4000:], parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    else:
        await msg.edit_text(text, parse_mode=ParseMode.MARKDOWN)
        await q.message.reply_text("✅ Natijalar tayyor!", reply_markup=kb)

async def _finish_mock(message, context):
    my = context.user_data["mock_yech"]
    conn = db()
    correct_answers = conn.execute(
        "SELECT correct_answer FROM mock_questions WHERE test_id=? ORDER BY order_num",
        (my["test_id"],)).fetchall()
    conn.close()
    score = sum(1 for i, ca in enumerate(correct_answers)
                if i < len(my["answers"]) and my["answers"][i].lower() == ca[0].lower())
    total = len(correct_answers)
    pct = (score/total*100) if total > 0 else 0
    conn = db()
    conn.execute(
        "INSERT INTO mock_results(test_id,student_name,score,total,percentage,answers)VALUES(?,?,?,?,?,?)",
        (my["test_id"], my["student_name"], score, total, pct, str(my["answers"])))
    conn.commit(); conn.close()
    if pct >= 90: baho = "A'lo ⭐⭐⭐"
    elif pct >= 75: baho = "Yaxshi ⭐⭐"
    elif pct >= 60: baho = "Qoniqarli ⭐"
    else: baho = "Qoniqarsiz 📚"
    bar = "█"*int(pct/10) + "░"*(10-int(pct/10))
    uid = message.chat.id
    kb = admin_kb() if is_admin(uid) else student_kb()
    # OTM topish tugmasi
    otm_kb = InlineKeyboardMarkup([[
        InlineKeyboardButton(
            f"🎯 {score*1.4:.0f} ball bilan qaysi OTMga kiraman?",
            callback_data=f"mock_to_dtm_{int(score*1.4)}"
        )
    ]])

    await message.reply_text(
        f"🎉 *Test yakunlandi!*\n\n"
        f"👤 {my['student_name']}\n📝 {my['test_title']}\n━━━━━━━━━━━━━━━\n"
        f"✅ To'g'ri: {score}/{total}\n📊 {pct:.0f}%\n|{bar}|\n🏆 {baho}\n\n"
        f"💡 _DTM ball taxminiy hisobi: {score*1.4:.0f} ball_",
        parse_mode=ParseMode.MARKDOWN, reply_markup=otm_kb)
    context.user_data.pop("mock_yech", None)

# ─── HISOBOT ───────────────────────────────────────────────────
async def admin_report_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Umumiy hisobot", callback_data="rep_general")],
        [InlineKeyboardButton("🎭 To'garaklar", callback_data="rep_clubs")],
        [InlineKeyboardButton("🏆 Yutuqlar", callback_data="rep_ach")],
    ])
    await update.message.reply_text("📊 *HISOBOT*", parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def cb_rep_general(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    ts = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    tc = conn.execute("SELECT COUNT(*) FROM clubs").fetchone()[0]
    ta = conn.execute("SELECT COUNT(*) FROM achievements").fetchone()[0]
    tm = conn.execute("SELECT COUNT(*) FROM club_members").fetchone()[0]
    treg = conn.execute("SELECT COUNT(*) FROM students WHERE telegram_id IS NOT NULL").fetchone()[0]
    classes = conn.execute("SELECT class_name,COUNT(*) FROM students GROUP BY class_name ORDER BY COUNT(*) DESC LIMIT 5").fetchall()
    top = conn.execute("SELECT s.full_name,COUNT(a.id) FROM students s JOIN achievements a ON s.id=a.student_id GROUP BY s.id ORDER BY COUNT(a.id) DESC LIMIT 5").fetchall()
    conn.close()
    data = (f"Jami o'quvchilar: {ts} (botda: {treg})\n"
            f"To'garaklar: {tc}, Yutuqlar: {ta}, A'zolar: {tm}\n"
            f"Sinflar: {classes}\nTop o'quvchilar: {top}")
    msg = await q.message.reply_text("⏳ Hisobot...")
    await msg.edit_text(ai_report("UMUMIY", data))
    return MAIN_MENU

async def cb_rep_clubs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    rows = conn.execute("SELECT cl.name,cl.direction,COUNT(cm.id) FROM clubs cl LEFT JOIN club_members cm ON cl.id=cm.club_id GROUP BY cl.id ORDER BY COUNT(cm.id) DESC").fetchall()
    conn.close()
    if not rows: await q.message.reply_text("📭 To'garaklar yo'q."); return MAIN_MENU
    data = "\n".join([f"{r[0]}: {r[1]}, {r[2]} a'zo" for r in rows])
    msg = await q.message.reply_text("⏳ Hisobot...")
    await msg.edit_text(ai_report("TO'GARAKLAR", data))
    return MAIN_MENU

async def cb_rep_ach(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    conn = db()
    rows = conn.execute("SELECT s.full_name,s.class_name,a.title,a.achievement_type,a.date FROM achievements a JOIN students s ON a.student_id=s.id ORDER BY a.date DESC").fetchall()
    conn.close()
    if not rows: await q.message.reply_text("📭 Yutuqlar yo'q."); return MAIN_MENU
    data = "\n".join([f"{r[0]}({r[1]}): {r[2]}, {r[3]}, {r[4]}" for r in rows])
    msg = await q.message.reply_text("⏳ Hisobot...")
    await msg.edit_text(ai_report("YUTUQLAR", data))
    return MAIN_MENU

# ─── KASB YO'NALTIRISH ─────────────────────────────────────────
async def career_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🎯 DTM Ball bo'yicha OTM topish", callback_data="dtm_start")],
        [InlineKeyboardButton("🏫 Barcha OTMlar", callback_data="otm_list")],
        [InlineKeyboardButton("❓ AI kasb maslahat", callback_data="teach_career")],
    ])
    await update.message.reply_text("🎯 *KASB YO'NALTIRISH*", parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    return MAIN_MENU

async def dtm_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    from dtm_base import FANLAR
    kb_rows = []
    row = []
    for i, fan in enumerate(FANLAR):
        row.append(InlineKeyboardButton(fan, callback_data=f"dtm_f1_{fan}"))
        if len(row) == 2: kb_rows.append(row); row = []
    if row: kb_rows.append(row)
    await q.message.reply_text(
        "🎯 *DTM ball bo'yicha OTM topish*\n\n1-fanni tanlang:",
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb_rows))
    return DTM_FAN1

async def dtm_fan1(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    fan1 = q.data.replace("dtm_f1_", "")
    context.user_data["dtm_fan1"] = fan1
    from dtm_base import FANLAR
    kb_rows = []; row = []
    for fan in FANLAR:
        if fan == fan1: continue
        row.append(InlineKeyboardButton(fan, callback_data=f"dtm_f2_{fan}"))
        if len(row) == 2: kb_rows.append(row); row = []
    if row: kb_rows.append(row)
    await q.message.reply_text(
        f"✅ 1-fan: *{fan1}*\n\n2-fanni tanlang:",
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb_rows))
    return DTM_FAN2

async def dtm_fan2(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data["dtm_fan2"] = q.data.replace("dtm_f2_", "")
    await q.message.reply_text(
        f"✅ 1-fan: *{context.user_data['dtm_fan1']}*\n"
        f"✅ 2-fan: *{context.user_data['dtm_fan2']}*\n\n"
        f"Jami ballingizni kiriting _(100-200)_:",
        parse_mode=ParseMode.MARKDOWN)
    return DTM_BALL

async def dtm_ball(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text in MENU_ITEMS: return await handle_menu(update, context)
    try:
        ball = float(update.message.text.strip().replace(",", "."))
        if not 50 <= ball <= 200: raise ValueError
    except:
        await update.message.reply_text("❗ Ball 50-200 oralig'ida bo'lishi kerak:")
        return DTM_BALL
    fan1 = context.user_data["dtm_fan1"]; fan2 = context.user_data["dtm_fan2"]
    msg = await update.message.reply_text("🔍 Qidirilmoqda...")
    from dtm_base import find_universities
    natija = find_universities(fan1, fan2, ball)
    grant_list = natija["grant"]; kontrakt_list = natija["kontrakt"]
    if not grant_list and not kontrakt_list:
        await msg.edit_text(
            f"😔 *{ball} ball* bilan mos yo'nalish topilmadi.\n\n"
            f"💡 Ball oshiring yoki boshqa fan kombinatsiyasini sinang.",
            parse_mode=ParseMode.MARKDOWN)
        return MAIN_MENU
    text = f"🎯 *{fan1} + {fan2} | {ball} ball*\n━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    if grant_list:
        text += f"🎓 *GRANT ({len(grant_list)} ta):*\n\n"
        for i, r in enumerate(grant_list, 1):
            text += (f"{i}. *{r['otm']}*\n   📚 {r['yonalish']}\n"
                     f"   🏙 {r['city']} | 🎓 {r['grant_ball']} ball | ⏱ {r['muddat']} yil\n\n")
    if kontrakt_list:
        text += f"💳 *KONTRAKT ({len(kontrakt_list)} ta):*\n\n"
        for i, r in enumerate(kontrakt_list, 1):
            text += (f"{i}. *{r['otm']}*\n   📚 {r['yonalish']}\n"
                     f"   🏙 {r['city']} | 💳 {r['kontrakt_ball']} ball | 💰 {r['narx']} mln/yil\n\n")
    kb = admin_kb() if is_admin(update.effective_user.id) else student_kb()
    if len(text) > 4000:
        await msg.edit_text(text[:4000], parse_mode=ParseMode.MARKDOWN)
        await update.message.reply_text(text[4000:], parse_mode=ParseMode.MARKDOWN, reply_markup=kb)
    else:
        await msg.edit_text(text, parse_mode=ParseMode.MARKDOWN)
        await update.message.reply_text("✅ Natijalar tayyor!", reply_markup=kb)
    return MAIN_MENU

async def otm_list_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    from dtm_base import DTM_BALLARI
    text = "🏫 *O'ZBEKISTON OTMlari:*\n\n"
    for otm_id, otm in DTM_BALLARI.items():
        mn = min(y["grant"] for y in otm["yonalishlar"])
        mx = max(y["grant"] for y in otm["yonalishlar"])
        text += f"🎓 *{otm['full_name']}*\n   🏙 {otm['city']} | 🎓 {mn}–{mx} ball\n\n"
    if len(text) > 4000:
        await q.message.reply_text(text[:4000], parse_mode=ParseMode.MARKDOWN)
        await q.message.reply_text(text[4000:], parse_mode=ParseMode.MARKDOWN)
    else:
        await q.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)
    return MAIN_MENU

# ─── EXCEL YUKLASH ─────────────────────────────────────────────
async def excel_upload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    doc = update.message.document
    if not doc or not (doc.file_name or "").endswith((".xlsx", ".xls")): return
    msg = await update.message.reply_text("📊 Excel o'qilmoqda...")
    try:
        file = await doc.get_file()
        file_bytes = await file.download_as_bytearray()
        wb = openpyxl.load_workbook(io.BytesIO(bytes(file_bytes)))
        ws = wb.active
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        name_idx = 0
        class_idx = None
        for i, h in enumerate(headers):
            if any(x in h for x in ["полное наим", "full_name", "ism", "f.i.o", "фамилия"]):
                name_idx = i
            if any(x in h for x in ["класс", "sinf", "class", "klass"]):
                class_idx = i
        conn = db(); c = conn.cursor()
        added = skipped = duplicates = 0
        last_row = 1
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            last_row = row_num
            if not row or not row[name_idx]: skipped += 1; continue
            name = str(row[name_idx]).strip()
            if len(name) < 3: skipped += 1; continue
            cls = ""
            if class_idx is not None and class_idx < len(row) and row[class_idx]:
                cls = str(row[class_idx]).strip()
            if not cls:
                # F/G ustunidan sinf qidirish
                for i, h in enumerate(headers):
                    if any(x in h for x in ["klass", "класс", "sinf"]) and i < len(row) and row[i]:
                        raw = str(row[i])
                        for part in raw.split():
                            if len(part) <= 4 and any(c2.isdigit() for c2 in part):
                                cls = part; break
                        break
            if conn.execute("SELECT id FROM students WHERE full_name=?", (name,)).fetchone():
                duplicates += 1; continue
            c.execute("INSERT INTO students(full_name,class_name) VALUES(?,?)", (name, cls))
            added += 1
        conn.commit(); conn.close()
        await msg.edit_text(
            f"✅ *Excel yuklandi!*\n\n"
            f"✅ Qo'shildi: *{added}* ta\n"
            f"🔁 Takror: {duplicates} ta\n"
            f"⚠️ Bo'sh: {skipped} ta",
            parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        await msg.edit_text(f"❌ Xato: {e}")

# ─── HANDLE MENU HELPER ────────────────────────────────────────
async def handle_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text
    context.user_data.pop("tsid", None)
    context.user_data.pop("tsname", None)
    context.user_data.pop("ai_mode", None)

    if is_admin(uid):
        routes = {
            "👨‍🎓 O'quvchilar": admin_students_menu,
            "🏆 Yutuqlar": admin_achievements_menu,
            "🎭 To'garaklar": admin_clubs_menu,
            "🎯 Kasb yo'naltirish": career_menu,
            "📋 Portfel": admin_portfolio_menu,
            "📝 MOCK test": admin_mock_menu,
            "📊 Hisobot": admin_report_menu,
            "📨 Murojaatlar": murojaatlar_admin,
            "📊 DTM Ballarni yangilash": dtm_yangilash_cmd,
        }
    else:
        routes = {
            "🎭 To'garaklar": student_clubs_menu,
            "🏆 Mening yutuqlarim": student_my_achievements,
            "📝 MOCK test yechish": student_mock_menu,
            "📨 Adminga murojaat": murojaat_start,
            "🎯 OTM topish (DTM ball)": student_dtm_menu,
            "👤 Mening profilim": student_profile,
        }
    if text in routes:
        result = await routes[text](update, context)
        return result if result else MAIN_MENU
    return MAIN_MENU

# ─── CALLBACK DISPATCHER ──────────────────────────────────────
async def cb_dispatch(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    d = q.data

    simple = {
        "add_student": cb_add_student,
        "srch_student": cb_srch_student,
        "list_students": cb_list_students,
        "add_ach": cb_add_ach,
        "list_ach": cb_list_ach,
        "add_club": cb_add_club,
        "list_clubs": cb_list_clubs,
        "club_members_list": cb_club_members,
        "dtm_start": dtm_start,
        "otm_list": otm_list_cb,
        "mock_create": mock_create,
        "mock_list": mock_list_cb,
        "mock_results_menu": mock_results_menu_cb,
        "mock_save": mock_save,
        "rep_general": cb_rep_general,
        "rep_clubs": cb_rep_clubs,
        "rep_ach": cb_rep_ach,
        "excel_help": lambda u,c: (
            u.callback_query.answer(),
            u.callback_query.message.reply_text(
                "📥 *Excel yuklash:*\n\nExcel (.xlsx) faylni shu chatga yuboring.\n"
                "Bot avtomatik barcha o'quvchilarni qo'shadi.\n\n"
                "*Ustunlar:* A=Ism, G=Sinf",
                parse_mode=ParseMode.MARKDOWN)
        )[1],
        "teach_career": lambda u,c: (
            u.callback_query.answer(),
            u.callback_query.message.reply_text(
                gemini("Kasb tanlashda o'quvchilarga qanday yordam berish kerak? Qisqa maslahat.",
                       "Siz maktab maslahatchisisiz. O'zbek tilida javob bering."))
        )[1],
    }

    for key, handler in simple.items():
        if d == key:
            return await handler(update, context)

    if d.startswith("cd_"): return await cb_cdir(update, context)
    if d.startswith("at_"): return await cb_atype(update, context)
    if d.startswith("sp_"): return await cb_student_profile(update, context)
    if d.startswith("asel_"): return await cb_asel(update, context)
    if d.startswith("msel_"): pass
    if d.startswith("mjoin_"): pass
    if d.startswith("mres_"): return await mock_result_show(update, context)
    if d.startswith("mq_"): return await mock_answer(update, context)
    if d.startswith("mqtype_"): return await mock_qtype(update, context)
    if d.startswith("mjavc_"): return await mock_correct(update, context)
    if d.startswith("mcls_"): return await mock_class(update, context)
    if d.startswith("dtm_f1_"): return await dtm_fan1(update, context)
    if d.startswith("dtm_f2_"): return await dtm_fan2(update, context)
    if d.startswith("cmemb_"): return await cb_show_members(update, context)
    if d.startswith("start_test_"): return await cb_start_test(update, context)
    if d.startswith("sjoin_"): return await student_join_club(update, context)
    if d.startswith("sleave_"): return await student_leave_club(update, context)
    if d.startswith("pf_"):
        sid = int(d.split("_")[-1]); await q.answer()
        conn = db()
        s = conn.execute("SELECT full_name,class_name FROM students WHERE id=?", (sid,)).fetchone()
        media = conn.execute("SELECT media_type,content,caption,added_at FROM student_media WHERE student_id=? ORDER BY added_at", (sid,)).fetchall()
        conn.close()
        if not s: return MAIN_MENU
        if not media:
            await q.message.reply_text(f"⚠️ Ma'lumot yo'q. /add_media {sid}"); return MAIN_MENU
        msg = await q.message.reply_text(f"⏳ Portfel...")
        data = [{"media_type":r[0],"content":r[1],"caption":r[2],"added_at":r[3]} for r in media]
        await msg.edit_text(ai_portfolio(s[0], s[1], data))
        return MAIN_MENU
    if d.startswith("am_"):
        sid = int(d.split("_")[-1]); await q.answer()
        row = db().execute("SELECT full_name FROM students WHERE id=?", (sid,)).fetchone()
        if row:
            context.user_data["tsid"] = sid; context.user_data["tsname"] = row[0]
            await q.message.reply_text(f"📁 *{row[0]}* uchun yuboring:\n📝 Matn | 🖼 Rasm | 🎥 Video\n\n✅ /done", parse_mode=ParseMode.MARKDOWN)
        return S_DATA
    if d.startswith("aa_"):
        sid = int(d.split("_")[-1]); await q.answer()
        row = db().execute("SELECT full_name FROM students WHERE id=?", (sid,)).fetchone()
        if row:
            context.user_data.update({"asid": sid, "asname": row[0]})
            await q.message.reply_text(f"🏆 *{row[0]}* yutuq nomi:", parse_mode=ParseMode.MARKDOWN)
        return A_TITLE

    await q.answer("⚙️")
    return MAIN_MENU

# ─── ASOSIY ROUTER ─────────────────────────────────────────────
async def main_router(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text or ""

    # Ro'yxatdan o'tmagan oddiy foydalanuvchi
    if not is_admin(uid) and not get_student_by_tg(uid):
        await update.message.reply_text("❗ Iltimos, /start bosing.")
        return MAIN_MENU

    # Menyu tugmasi
    if text in MENU_ITEMS:
        return await handle_menu(update, context)

    # Media qo'shish rejimi (admin)
    if context.user_data.get("tsid"):
        return await receive_media(update, context)

    # DTM yangilash rejimi
    if context.user_data.get("dtu_step") and is_admin(uid):
        await dtu_input_handler(update, context)
        return MAIN_MENU

    # AI rejim
    if context.user_data.get("ai_mode"):
        msg = await update.message.reply_text("🤔 AI o'ylamoqda...")
        answer = ai_guide(text)
        await msg.edit_text(f"🤖 *AI:*\n\n{answer}", parse_mode=ParseMode.MARKDOWN)
        return MAIN_MENU

    return MAIN_MENU

# ─── MAIN ──────────────────────────────────────────────────────
def _load_dtm_updates():
    pass
    
async def main():
    init_db()
    _load_dtm_updates()  # Saqlangan DTM yangilanishlarni yuklash
    logger.info("🚀 Bot v6 ishga tushmoqda...")
    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MAIN_MENU: [
                MessageHandler(filters.ALL & ~filters.COMMAND, main_router),
                CallbackQueryHandler(cb_dispatch),
            ],
            REG_NAME:       [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_name)],
            REG_CLASS:      [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_class)],
            S_NAME:         [MessageHandler(filters.TEXT & ~filters.COMMAND, s_name)],
            S_CLASS:        [MessageHandler(filters.TEXT & ~filters.COMMAND, s_class)],
            S_DATA: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_media),
                MessageHandler(filters.PHOTO, receive_media),
                MessageHandler(filters.VIDEO, receive_media),
                MessageHandler(filters.Document.ALL, receive_media),
                CommandHandler("done", done_cmd),
            ],
            S_SEARCH:       [MessageHandler(filters.TEXT & ~filters.COMMAND, s_search)],
            PORTFOLIO:      [MessageHandler(filters.TEXT & ~filters.COMMAND, portfolio_handler)],
            C_NAME:         [MessageHandler(filters.TEXT & ~filters.COMMAND, c_name)],
            C_DIR:          [CallbackQueryHandler(cb_cdir, pattern="^cd_")],
            C_RESP:         [MessageHandler(filters.TEXT & ~filters.COMMAND, c_resp)],
            A_STUDENT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, a_student),
                CallbackQueryHandler(cb_asel, pattern="^asel_"),
            ],
            A_TITLE:        [MessageHandler(filters.TEXT & ~filters.COMMAND, a_title)],
            A_TYPE:         [CallbackQueryHandler(cb_atype, pattern="^at_")],
            A_DATE:         [MessageHandler(filters.TEXT & ~filters.COMMAND, a_date)],
            AI_CHAT:        [MessageHandler(filters.TEXT & ~filters.COMMAND, main_router)],
            M_SEARCH:       [MessageHandler(filters.TEXT & ~filters.COMMAND, main_router)],
            M_CLUB:         [CallbackQueryHandler(cb_dispatch)],
            DTM_FAN1:       [CallbackQueryHandler(dtm_fan1, pattern="^dtm_f1_")],
            DTM_FAN2:       [CallbackQueryHandler(dtm_fan2, pattern="^dtm_f2_")],
            DTM_BALL:       [MessageHandler(filters.TEXT & ~filters.COMMAND, dtm_ball)],
            MOCK_NEW_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, mock_title)],
            MOCK_NEW_CLASS: [CallbackQueryHandler(mock_class, pattern="^mcls_")],
            MOCK_NEW_FAN:   [MessageHandler(filters.TEXT & ~filters.COMMAND, mock_fan)],
            MOCK_SAVOL: [
                CallbackQueryHandler(mock_qtype, pattern="^mqtype_"),
                CallbackQueryHandler(mock_save, pattern="^mock_save"),
            ],
            MOCK_JAVOB_TEXT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, mock_qtext),
                CallbackQueryHandler(mock_correct, pattern="^mjavc_"),
            ],
            MOCK_YECHISH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, main_router),
                CallbackQueryHandler(mock_answer, pattern="^mq_"),
            ],
            MUROJAAT_TEXT:  [MessageHandler(filters.TEXT & ~filters.COMMAND, murojaat_text)],
        },
        fallbacks=[
            CommandHandler("start", start),
        ],
        allow_reentry=True,
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("add_media", add_media_cmd))
    app.add_handler(CommandHandler("add_togarak", add_togarak_cmd))
    app.add_handler(CommandHandler("mock_yech", mock_yech_cmd))
   app.add_handler(CommandHandler("dtm_yangilash", lambda u,c: dtm_yangilash_start(u,c)))
    app.add_handler(MessageHandler(
        filters.Document.ALL & filters.ChatType.PRIVATE, excel_upload_handler))
    app.add_handler(CallbackQueryHandler(cb_dispatch))

    logger.info("✅ Bot v6 tayyor!")
    async with app:
        await app.initialize()
        await app.start()
        await app.updater.start_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)
        logger.info("✅ Bot ishlayapti!")
        await asyncio.Event().wait()

if __name__ == "__main__":
    asyncio.run(main())
